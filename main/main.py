
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
from PIL import ImageTk as imtk
from tkinter import filedialog
from datetime import datetime
from PIL import Image as im
from datetime import date
from tkinter import ttk
from tkinter import *
import customtkinter
import pandas as pd
import re

#==============================================================================

def run():

    global path

    try:
        path = filedialog.askopenfilename(title='選取檔案', multiple=False)
        main()

    except:
        text='檔案無法辨識'
        label_info.configure(text = text)

        for value in tree.get_children():
            tree.delete(value)

#------------------------------------------------------------------------------

def main():
    
    global data

    global path

    wb = load_workbook(path)
    sheet = wb.active

    ID_list   = []
    name_list = []
    sex_list  = []
    date_list = []

    index = 2
    while True:
        if sheet.cell(row = index, column = 1).value == None and sheet.cell(row = index, column = 3).value == None and sheet.cell(row = index, column = 4).value == None and sheet.cell(row = index, column = 5).value == None:
            break
        ID_list.append(sheet.cell(row = index, column = 1).value)
        name_list.append(sheet.cell(row = index, column = 3).value)
        sex_list.append(sheet.cell(row = index, column = 4).value)
        date_list.append(sheet.cell(row = index, column = 5).value)
        index += 1

    dict = {'身分證字號': ID_list[1:], '姓名': name_list[1:], '性別': sex_list[1:], '生日': date_list[1:]}
    data = pd.DataFrame(dict).astype('str')

    data[''] = data.apply(chack_sapce, axis=1)
    data[''] = data.apply(check_char, axis=1)
    data[''] = data.apply(check_question, axis=1)
    data[['', '更正後性別']] = data.apply(chack_ID, axis=1)
    data[['', '更正後生日']] = data.apply(chack_date, axis=1)
    data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] = data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] + ' (身份證字號重複)'

    data = data.reindex(columns=['身分證字號', '姓名', '性別', '更正後性別', '生日', '更正後生日', ''])

#------------------------------------------------------------------------------

    savepath = path[0:-len(path.split('/')[-1])]
    data[data[''] != ''].to_excel(savepath+'檢查結果.xlsx', index=False)

    text='有問題數量: '+str(data[data[''] != ''].shape[0])
    label_info.configure(text = text)

    for value in tree.get_children():
        tree.delete(value)

    for value in data[data[''] != ''].values.tolist():
        tree.insert('', END, values=value)

#------------------------------------------------------------------------------

def save():
    
    global data
    
    global path
    
    save_data = data
    save_data['性別'] = save_data['更正後性別']
    save_data['生日'] = save_data['更正後生日']
    
    save_data[''] = save_data.apply(chack_sapce, axis=1)
    save_data[''] = save_data.apply(check_char, axis=1)
    save_data[''] = save_data.apply(check_question, axis=1)
    save_data[['', '更正後性別']] = save_data.apply(chack_ID, axis=1)
    save_data[['', '更正後生日']] = save_data.apply(chack_date, axis=1)
    save_data.loc[save_data.duplicated(subset=['身分證字號'], keep=False), ''] = save_data.loc[save_data.duplicated(subset=['身分證字號'], keep=False), ''] + ' (身份證字號重複)'
    
    output_wb = load_workbook('file/勞工報告表格範例.xlsx')
    output_sheet = output_wb.active
    for index, value in save_data.iterrows():
        output_sheet.cell(row=index+3, column=1).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=1).value = value['身分證字號']

        output_sheet.cell(row=index+3, column=3).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=3).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=3).value = value['姓名']

        output_sheet.cell(row=index+3, column=4).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=4).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=4).value = value['更正後性別']

        output_sheet.cell(row=index+3, column=5).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=5).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=5).value = value['更正後生日']

        output_sheet.cell(row=index+3, column=19).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=19).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=19).value = value['']
    
    savepath = path[0:-len(path.split('/')[-1])]
    output_wb.save(savepath+'勞工報告表格.xlsx')
    
    message_box('勞工報告表格已輸出')

#------------------------------------------------------------------------------

def chack_sapce(data_in):

    temp = ''
    for name in ['身分證字號', '姓名', '性別', '生日']:
        if pd.isna(data_in[name]) or data_in[name] == 'None' or data_in[name] == '':
            temp += ' (' + name + '資料缺失)'

        else:
            data_in[name] = re.sub(' +', ' ', data_in[name])
            
            for index, char in enumerate(data_in[name]):
                if char == ' ':
                    if index == 0:
                        temp += ' (' + name + '中的' + data_in[name][1] + '前有空格)'

                    elif index == len(data_in[name])-1:
                        temp += ' (' + name + '中的' + data_in[name][index-1] + '後有空格)'

                    else:
                        temp += ' (' + name + '中的' + data_in[name][index-1]  + '與' + data_in[name][index+1] + '中間有空格)'

    return temp

#------------------------------------------------------------------------------

def chack_ID(data_in):

    code_list = ['10','11','12','13','14','15','16','17','34','18','19','20','21','22','35','23','24','25','26','27','28','29','32','30','31','33']
    
    weight_list = [1,9,8,7,6,5,4,3,2,1,1]
    
    data_in['身分證字號'] = data_in['身分證字號'].replace(" ", "")
    data_in['性別'] = data_in['性別'].replace(" ", "")

    try:
        newcode = [int(x) for x in str(code_list[ord(data_in['身分證字號'][0])-65] + data_in['身分證字號'][1:])]

        if newcode[2] not in [1,2]:
            return pd.Series([data_in[''] + ' (外籍人士請留證件)', data_in['性別']])

        else:
            num = 0
            for code_value, weight_value in zip(newcode, weight_list):
                num += code_value  * weight_value
        
            if num % 10 == 0 and len(data_in['身分證字號'][1:]) == 9:

                if data_in['身分證字號'][1] == '1' and data_in['性別'] not in ['男','M','m','MALE','Male','male']:
                    return pd.Series([data_in[''] + ' (性別有誤請留證件)', '男'])

                elif data_in['身分證字號'][1] == '2' and data_in['性別'] not in ['女','F','f','FEMALE','Female','female']:
                    return pd.Series([data_in[''] + ' (性別有誤請留證件)', '女'])

                else:
                    if data_in['身分證字號'][1] == '1' and data_in['性別'] != '男':
                        return pd.Series([data_in[''] + ' (性別有誤請留證件)', '男'])
    
                    elif data_in['身分證字號'][1] == '2' and data_in['性別'] != '女':
                        return pd.Series([data_in[''] + ' (性別有誤請留證件)', '女'])
                    
                    else:
                        return pd.Series([data_in[''], data_in['性別']])

            else:
                return pd.Series([data_in[''] + ' (身分證有誤請留證件)', data_in['性別']])
        
    except:
        if not data_in['身分證字號'][1].isnumeric():
            return pd.Series([data_in[''] + ' (外籍人士請留證件)', data_in['性別']])

        else:
            return pd.Series([data_in[''] + ' (身分證異常請留證件)', data_in['性別']])

#------------------------------------------------------------------------------

def check_char(data_in):

    data_in['姓名'] = data_in['姓名'].replace(" ", "")

    with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
        word = f.read().split(',')

    temp = []
    for char in [x for x in data_in['姓名']]:
        if char in word:
            temp.append(char)

    if len(temp) > 0:
        return data_in[''] + ' (姓名為難字請留證件)'

    else:
        return data_in['']

#------------------------------------------------------------------------------

def check_question(data_in):

    data_in['姓名'] = data_in['姓名'].replace(" ", "")

    question = False
    for char in [x for x in data_in['姓名']]:
        if char == '?':
            question = True

    if question:
        return data_in[''] + ' (姓名異常請留證件)'

    else:
        return data_in['']

#------------------------------------------------------------------------------

def chack_date(data_in):

    data_in['生日'] = data_in['生日'].replace(" ", "")
    
    try:
        newdate = datetime.strptime(data_in['生日'], '%Y/%m/%d').strftime('%Y/%m/%d')
        
        if data_in['生日'] == newdate:
            return pd.Series([data_in[''], data_in['生日']])
        
        else:
            return pd.Series([data_in[''] + ' (生日有誤)', newdate])
            
    except:
        try:

            date_str = re.split(r'\D+', data_in['生日'])[0:3]

            if int(date_str[0][0]) == 0:
                date_str[0] = date_str[0][1:]

            if int(date_str[0]) < 1911:
                date_str[0] = str(int(date_str[0])+1911)

            date_str[1] = date_str[1].zfill(2)
            date_str[2] = date_str[2].zfill(2)

            date_str = '/'.join(date_str)

            newdate = datetime.strptime(date_str, '%Y/%m/%d').strftime('%Y/%m/%d')
            return pd.Series([data_in[''] + ' (生日有誤)', newdate])

        except:
            try:
                if (date.today().year-65) < int(data_in['生日'][0:4]) <= date.today().year:
                    newdate = datetime.strptime(data_in['生日'], '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

                elif (date.today().year-65-1911) < int(data_in['生日'][0:2]) <= (date.today().year-1911):
                    newdate = datetime.strptime(str(int(data_in['生日']) + 1911*10**(len(str(data_in['生日']))-2)), '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

                elif (date.today().year-65-1911) < int(data_in['生日'][0:3]) <= (date.today().year-1911):
                    newdate = datetime.strptime(str(int(data_in['生日']) + 1911*10**(len(str(data_in['生日']))-3)), '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

            except:
                return pd.Series([data_in[''] + ' (生日有誤無法更正)', ''])

#------------------------------------------------------------------------------

def load_image(impath, image_size):

    return imtk.PhotoImage(im.open(impath).resize((image_size, image_size)))

#------------------------------------------------------------------------------

def add_char():

    add = str(char_entry.get())

    if add == '':
        message_box('未輸入陽明難字')

    else:
        with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
            word = f.read().split(',')

        for value in add.split(','):
            if value not in word:
                word.append(value)

        with open('file/word.txt', 'w', encoding='utf_8_sig') as f:
            f.write(','.join(word))

        message_box('陽明難字已新增: '+add)

#------------------------------------------------------------------------------

def minus_char():

    minus = str(char_entry.get())

    if minus == '':
        message_box('未輸入陽明難字')

    else:
        with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
            word = f.read().split(',')

        for value in minus.split(','):
            if value in word:
                word.remove(value)

        with open('file/word.txt', 'w', encoding='utf_8_sig') as f:
            f.write(','.join(word))

        message_box('陽明難字已移除: '+minus)

#------------------------------------------------------------------------------

def message_box(text):
    
    message_font = ('Courier New','10')
    
    message = Toplevel()
    message.title('Message')
    message.geometry('250x100')
    message.config(bg='#292929')
    
    message_frame = customtkinter.CTkFrame(master=message, corner_radius=10)
    message_frame.place(relx=0.5, rely=0.5, relwidth=0.95, relheight=0.95, anchor=CENTER)

    message_Label = Label(message_frame,text=text, font=message_font, fg='silver')
    message_Label.config(bg='#292929')
    message_Label.place(relx=0.5,rely=0.4,anchor='center')
    
    close_message_button = customtkinter.CTkButton(master=message_frame, text='OK', command=message.destroy)
    close_message_button.place(relx=0.5,rely=0.6, anchor='n')

#==============================================================================

customtkinter.set_appearance_mode('dark')
customtkinter.set_default_color_theme('blue')
text_font = ('Courier New','10')

app = customtkinter.CTk()
app.geometry('1000x520')
app.title('Hsu Mei')
app.resizable(0,0) 
app.iconbitmap('icon/HsuMei_icon.ico')
app.grid_columnconfigure(0, weight=10)
app.grid_columnconfigure(1, weight=1)
app.grid_rowconfigure(0, weight=1)

#-Style------------------------------------------------------------------------

app_style = ttk.Style()
app_style.theme_use('clam')
app_style.configure('Treeview', foreground='white', background='gray38', fieldbackground='gray38')
app_style.configure('Treeview.Heading', foreground='white', background='gray38', font=text_font)
app_style.map('Treeview', background=[('selected', 'black')])

#-frame_left-------------------------------------------------------------------

frame_left = customtkinter.CTkFrame(master=app, corner_radius=10)
frame_left.grid(row=0, column=0, sticky=NSEW, padx=10, pady=10)
frame_left.grid_columnconfigure(0, weight=1)
frame_left.grid_rowconfigure(0, weight=1)

frame_treeview = customtkinter.CTkFrame(master=frame_left, corner_radius=0)
frame_treeview.grid(row=0, column=0, sticky=NSEW, padx=8, pady=8)
frame_treeview.grid_columnconfigure(0, weight=1)
frame_treeview.grid_rowconfigure(0, weight=1)

tree_scroll_Y = Scrollbar(frame_treeview, orient=VERTICAL)
tree_scroll_Y.grid(row=0, column=1, sticky=NSEW, padx=1, pady=1)

tree = ttk.Treeview(frame_treeview, 
                    columns=('ID', 'name', 'sex', 'newsex', 'date', 'newdate', 'error'), 
                    yscrollcommand=tree_scroll_Y.set, 
                    show='headings')

tree.heading('ID', text='身分證字號')
tree.heading('name', text='姓名')
tree.heading('sex', text='性別')
tree.heading('newsex', text='更正後性別')
tree.heading('date', text='生日')
tree.heading('newdate', text='更正後生日')
tree.heading('error', text='')

tree.column('ID', width=90, anchor=CENTER)
tree.column('name', width=60, anchor=CENTER)
tree.column('sex', width=40, anchor=CENTER)
tree.column('newsex', width=90, anchor=CENTER)
tree.column('date', width=90, anchor=CENTER)
tree.column('newdate', width=90, anchor=CENTER)
tree.column('error', anchor=W)

tree.grid(row=0, column=0, sticky=NSEW, padx=1, pady=1)

tree_scroll_Y.config(command=tree.yview)

#-frame_right------------------------------------------------------------------

frame_right = customtkinter.CTkFrame(master=app, corner_radius=0)
frame_right.grid(row=0, column=1, sticky=NSEW, padx=0, pady=0)
frame_right.grid_columnconfigure(0, weight=1)

label_logo = customtkinter.CTkLabel(master=frame_right,
                                    text='Hsu Mei',
                                    height=50,
                                    font=('Comic Sans MS', 26))
label_logo.grid(row=0, column=0, sticky=NSEW, padx=10, pady=20)

#-frame_right-up---------------------------------------------------------------

frame_right_up = customtkinter.CTkFrame(master=frame_right, corner_radius=10)
frame_right_up.grid(row=1, column=0, sticky=EW, padx=5, pady=5)
frame_right_up.grid_columnconfigure(0, weight=1)

label_info = customtkinter.CTkLabel(master=frame_right_up,
                                    text='',
                                    corner_radius=10,
                                    fg_color=('white', 'gray38'))
label_info.grid(row=1, column=0, sticky=EW, padx=10, pady=5)

select_image = load_image('icon/xls-file.png', 35)
select_button = customtkinter.CTkButton(master=frame_right_up,
                                        text='選取檔案',
                                        corner_radius=10,
                                        height=50,
                                        image=select_image,
                                        compound='right',
                                        command=run)
select_button.grid(row=2, column=0, sticky=EW, padx=10, pady=5)

refresh_image = load_image('icon/refresh.png', 35)
refresh_button = customtkinter.CTkButton(master=frame_right_up,
                                        text='重新整理',
                                        corner_radius=10,
                                        height=50,
                                        image=refresh_image,
                                        compound='right',
                                        command=main)
refresh_button.grid(row=3, column=0, sticky=EW, padx=10, pady=5)

save_image = load_image('icon/floppy-disk.png', 35)
save_button = customtkinter.CTkButton(master=frame_right_up,
                                        text='輸出檔案',
                                        corner_radius=10,
                                        height=50,
                                        image=save_image,
                                        compound='right',
                                        command=save)
save_button.grid(row=4, column=0, sticky=EW, padx=10, pady=5)

#-frame_right-down-------------------------------------------------------------

frame_right_down = customtkinter.CTkFrame(master=frame_right, corner_radius=10)
frame_right_down.grid(row=2, column=0, sticky=EW, padx=5, pady=5)
frame_right_down.grid_columnconfigure(0, weight=1)

char_entry = customtkinter.CTkEntry(master=frame_right_down, 
                                    placeholder_text='輸入難字',
                                    border_width=2,
                                    corner_radius=10)
char_entry.grid(row=0, column=0, sticky=EW, padx=10, pady=10)

add_image = load_image('icon/add.png', 25)
add_button = customtkinter.CTkButton(master=frame_right_down,
                                        text='新增難字',
                                        corner_radius=10,
                                        height=50,
                                        hover_color='green',
                                        image=add_image,
                                        compound='right',
                                        command=add_char)
add_button.grid(row=1, column=0, sticky=EW, padx=10, pady=5)

minus_image = load_image('icon/minus.png', 25)
minus_button = customtkinter.CTkButton(master=frame_right_down,
                                        text='刪除難字',
                                        corner_radius=10,
                                        height=50,
                                        hover_color='red',
                                        image=minus_image,
                                        compound='right',
                                        command=minus_char)
minus_button.grid(row=2, column=0, sticky=EW, padx=10, pady=5)

#------------------------------------------------------------------------------

app.mainloop()
