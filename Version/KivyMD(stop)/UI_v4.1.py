
from kivymd.uix.navigationdrawer import MDNavigationDrawerMenu
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.datatables import MDDataTable
from kivymd.toast import toast
from kivymd.app import MDApp

#------------------------------------------------------------------------------

from kivy.properties import ObjectProperty
from kivy.core.text import LabelBase
from kivy.core.window import Window
from kivy.config import Config
from kivy.lang import Builder
from kivy.metrics import dp

#------------------------------------------------------------------------------

from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from datetime import date
import pandas as pd
import winreg
import re

#==============================================================================

Window.size = (1000, 520)
LabelBase.register(name="Roboto", fn_regular="file/NaniFont-Regular.ttf")

HsuMeiKV = '''

MDScreen:

    MDTopAppBar:
        pos_hint: {"top": 1}
        elevation: 4
        title: "Hsu Mei"
        md_bg_color: "#BB5E00"
        right_action_items: 
            [
            ["file-upload", lambda x: app.file_manager_open(), "選擇人事資料", "選擇人事資料"],
            ["file-download", lambda x: app.export_result_file(), "輸出檢查結果", "輸出檢查結果"],
            ["file-export", lambda x: app.export_done_file(), "輸出更正結果", "輸出更正結果"],
            ["format-letter-matches", lambda x: app.to_scr2(), "陽明難字增減", "陽明難字增減"],
            ["reload", lambda x: app.reload(), "重新整理", "重新整理"]
            ]

    MDNavigationLayout:

        MDScreenManager:
            id: screen_manager

            MDScreen:
                name: "scr 1"
                id: main_layout

            MDScreen:
                name: "scr 2"
                MDTextField:
                    id: Textchar
                    pos_hint:{'center_x': 0.5, 'center_y': 0.6}
                    hint_text: "輸入難字"
                    required: True
                    helper_text_mode: "on_error"
                    helper_text: "未輸入陽明難字"
                    size_hint_x: 0.75

                MDRoundFlatIconButton:
                    pos_hint:{'center_x': 0.25, 'center_y': 0.4}
                    text: "新增難字"
                    icon: "plus"
                    text_color: "white"
                    on_press:
                        app.add_new_char()

                MDRoundFlatIconButton:
                    pos_hint:{'center_x': 0.5, 'center_y': 0.4}
                    text: "刪除難字"
                    icon: "minus"
                    text_color: "white"
                    on_press:
                        app.minus_old_char()

                MDRoundFlatIconButton:
                    pos_hint:{'center_x': 0.75, 'center_y': 0.4}
                    text: "返回"
                    icon: "keyboard-return"
                    text_color: "white"
                    on_press:
                        app.to_scr1()

'''

#==============================================================================

class HsuMei(MDApp):

    def __init__(self, **kwargs):

        super().__init__(**kwargs)
        self.manager_open = False
        self.file_manager = MDFileManager(exit_manager=self.exit_manager,
                                          select_path=self.select_path,
                                          background_color_toolbar="#BB5E00",
                                          background_color_selection_button="#BB5E00")

#------------------------------------------------------------------------------

    def build(self):

        self.icon = 'file/HsuMei_icon.ico'

        self.theme_cls.primary_palette = "Gray"
        self.theme_cls.theme_style = "Dark"

        self.screen = Builder.load_string(HsuMeiKV)

        self.data_tables = MDDataTable(pos_hint={"center_x": 0.5},
                                       size_hint=(1, 0.85),
                                       rows_num=100,
                                       use_pagination=True,
                                       column_data=[("[size=18]身分證字號", dp(27)),
                                                    ("[size=18]姓名", dp(13)),
                                                    ("[size=18]性別", dp(13)),
                                                    ("[size=18]更正後性別", dp(27)),
                                                    ("[size=18]生日", dp(27)),
                                                    ("[size=18]更正後生日", dp(27)),
                                                    ("[size=18]備註", dp(59))],)

        self.screen.ids.main_layout.add_widget(self.data_tables)

        return self.screen

#------------------------------------------------------------------------------

    def file_manager_open(self):

        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Shell Folders')
        self.file_manager.show(winreg.QueryValueEx(key, "Desktop")[0])
        self.manager_open = True

#------------------------------------------------------------------------------

    def select_path(self, path):

        self.path = path
        self.exit_manager()
        toast(self.path)

        Data = main(self.path)
        self.data_tables.row_data = Data.values

#------------------------------------------------------------------------------

    def exit_manager(self, *args):

        self.manager_open = False
        self.file_manager.close()

#------------------------------------------------------------------------------

    def export_result_file(self):

        try:
            save_result(self.path)
            toast('檢查結果已輸出')

        except:
            toast('未選擇人事資料')

#------------------------------------------------------------------------------

    def export_done_file(self):

        try:
            save_done(self.path)
            toast('勞工報告表格已輸出')

        except:
            toast('未選擇人事資料')

#------------------------------------------------------------------------------

    def reload(self):

        self.to_scr1()

        try:
            Data = main(self.path)
            self.data_tables.row_data = Data.values
            toast('已重新整理')

        except:
            toast('未選擇人事資料')

#------------------------------------------------------------------------------

    def to_scr1(self):

        self.screen.ids.screen_manager.transition.direction = 'right'
        self.screen.ids.screen_manager.current = 'scr 1'

#------------------------------------------------------------------------------

    def to_scr2(self):

        self.screen.ids.screen_manager.transition.direction = 'left'
        self.screen.ids.screen_manager.current = 'scr 2'

#------------------------------------------------------------------------------

    def add_new_char(self):
  
        if self.screen.ids.Textchar.text == '':
            toast('未輸入陽明難字')

        else:
            add_char(self.screen.ids.Textchar.text)
            toast('陽明難字已新增')

#------------------------------------------------------------------------------

    def minus_old_char(self):

        if self.screen.ids.Textchar.text == '':
            toast('未輸入陽明難字')

        else:
            minus_char(self.screen.ids.Textchar.text)
            toast('陽明難字已移除')

#==============================================================================

def main(path):

    global data

    wb = load_workbook(path)
    sheet = wb.active

    ID_list   = []
    name_list = []
    sex_list  = []
    date_list = []

    index = 2
    while True:
        if sheet.cell(row = index, column = 1).value == None:
            break
        ID_list.append(sheet.cell(row = index, column = 1).value)
        name_list.append(sheet.cell(row = index, column = 3).value)
        sex_list.append(sheet.cell(row = index, column = 4).value)
        date_list.append(sheet.cell(row = index, column = 5).value)
        index += 1

    dict = {'身分證字號': ID_list[1:], '姓名': name_list[1:], '性別': sex_list[1:], '生日': date_list[1:]}
    data = pd.DataFrame(dict).astype('str')

    data[''] = data.apply(chack_sapce, axis=1)
    data[''] = data.apply(check_char, axis=1)
    data[''] = data.apply(check_question, axis=1)
    data[['', '更正後性別']] = data.apply(chack_ID, axis=1)
    data[['', '更正後生日']] = data.apply(chack_date, axis=1)
    data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] = data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] + ' (身份證字號重複)'

    data = data.reindex(columns=['身分證字號', '姓名', '性別', '更正後性別', '生日', '更正後生日', ''])
    
    return data[data[''] != '']

#------------------------------------------------------------------------------

def save_result(path):

    global data

    savepath = path[0:-len(path.split('/')[-1])]
    if savepath == '':
        savepath = path[0:-len(path.split('\\')[-1])]
    data[data[''] != ''].to_excel(savepath+'檢查結果.xlsx', index=False)

#------------------------------------------------------------------------------

def save_done(path):
    
    global data
    
    save_data = data
    save_data['性別'] = save_data['更正後性別']
    save_data['生日'] = save_data['更正後生日']
    
    save_data[''] = save_data.apply(chack_sapce, axis=1)
    save_data[''] = save_data.apply(check_char, axis=1)
    save_data[''] = save_data.apply(check_question, axis=1)
    save_data[['', '更正後性別']] = save_data.apply(chack_ID, axis=1)
    save_data[['', '更正後生日']] = save_data.apply(chack_date, axis=1)
    save_data.loc[save_data.duplicated(subset=['身分證字號'], keep=False), ''] = save_data.loc[save_data.duplicated(subset=['身分證字號'], keep=False), ''] + ' (身份證字號重複)'
    
    output_wb = load_workbook('file/勞工報告表格範例.xlsx')
    output_sheet = output_wb.active
    for index, value in save_data.iterrows():
        output_sheet.cell(row=index+3, column=1).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=1).value = value['身分證字號']

        output_sheet.cell(row=index+3, column=3).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=3).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=3).value = value['姓名']

        output_sheet.cell(row=index+3, column=4).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=4).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=4).value = value['更正後性別']

        output_sheet.cell(row=index+3, column=5).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=5).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=5).value = value['更正後生日']

        output_sheet.cell(row=index+3, column=19).font = Font(size=10, name='新細明體')
        output_sheet.cell(row=index+3, column=19).alignment = Alignment(horizontal='center', vertical='center')
        output_sheet.cell(row=index+3, column=19).value = value['']
    
    savepath = path[0:-len(path.split('/')[-1])]
    if savepath == '':
        savepath = path[0:-len(path.split('\\')[-1])]
    output_wb.save(savepath+'勞工報告表格.xlsx')

#------------------------------------------------------------------------------

def chack_sapce(data_in):
    
    temp = ''
    for name in ['身分證字號', '姓名', '性別', '生日']:
        if pd.isna(data_in[name]) or data_in[name] == 'None' or data_in[name] == '':
            temp += ' (' + name + '資料缺失)'

        else:
            if data_in[name][0] == ' ' and data_in[name][-1] == ' ':
                temp += ' (' + name + '前後有空格)'
            
            elif data_in[name][-1] == ' ':
                temp += ' (' + name + '後有空格)'
            
            elif data_in[name][0] == ' ':
                temp += ' (' + name + '前有空格)'
            
    return temp

#------------------------------------------------------------------------------

def chack_ID(data_in):

    code_list = ['10','11','12','13','14','15','16','17','34','18','19','20','21','22','35','23','24','25','26','27','28','29','32','30','31','33']
    
    weight_list = [1,9,8,7,6,5,4,3,2,1,1]
    
    data_in['身分證字號'] = data_in['身分證字號'].strip()
    data_in['性別'] = data_in['性別'].strip()

    try:
        newcode = [int(x) for x in str(code_list[ord(data_in['身分證字號'][0])-65] + data_in['身分證字號'][1:])]

        if newcode[2] not in [1,2]:
            return pd.Series([data_in[''] + ' (外籍人士請留證件)', data_in['性別']])

        else:
            
            num = 0
            for code_value, weight_value in zip(newcode, weight_list):
                num += code_value  * weight_value
        
            if num % 10 == 0:

                if data_in['身分證字號'][1] == '1' and data_in['性別'] not in ['男','M','m','MALE','Male','male']:
                    return pd.Series([data_in[''] + ' (性別有誤請留證件)', '男'])

                elif data_in['身分證字號'][1] == '2' and data_in['性別'] not in ['女','F','f','FEMALE','Female','female']:
                    return pd.Series([data_in[''] + ' (性別有誤請留證件)', '女'])

                else:
                    if data_in['身分證字號'][1] == '1' and data_in['性別'] != '男':
                        return pd.Series([data_in[''] + ' (性別有誤請留證件)', '男'])
    
                    elif data_in['身分證字號'][1] == '2' and data_in['性別'] != '女':
                        return pd.Series([data_in[''] + ' (性別有誤請留證件)', '女'])
                    
                    else:
                        return pd.Series([data_in[''], data_in['性別']])

            else:
                return pd.Series([data_in[''] + ' (身分證有誤請留證件)', data_in['性別']])
        
    except:
        if not data_in['身分證字號'][1].isnumeric():
            return pd.Series([data_in[''] + ' (外籍人士請留證件)', data_in['性別']])

        else:
            return pd.Series([data_in[''] + ' (身分證異常請留證件)', data_in['性別']])

#------------------------------------------------------------------------------

def check_char(data_in):

    data_in['姓名'] = data_in['姓名'].strip()

    with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
        word = f.read().split(',')

    temp = []
    for char in [x for x in data_in['姓名']]:
        if char in word:
            temp.append(char)

    if len(temp) > 0:
        return data_in[''] + ' (姓名為難字請留證件)'

    else:
        return data_in['']

#------------------------------------------------------------------------------

def check_question(data_in):

    data_in['姓名'] = data_in['姓名'].strip()

    question = False
    for char in [x for x in data_in['姓名']]:
        if char == '?':
            question = True

    if question:
        return data_in[''] + ' (姓名異常請留證件)'

    else:
        return data_in['']

#------------------------------------------------------------------------------

def chack_date(data_in):

    data_in['生日'] = data_in['生日'].strip()
    
    try:
        newdate = datetime.strptime(data_in['生日'], '%Y/%m/%d').strftime('%Y/%m/%d')
        
        if data_in['生日'] == newdate:
            return pd.Series([data_in[''], data_in['生日']])
        
        else:
            return pd.Series([data_in[''] + ' (生日有誤)', newdate])
            
    except:
        try:

            date_str = re.split(r'\D+', data_in['生日'])[0:3]

            if int(date_str[0][0]) == 0:
                date_str[0] = date_str[0][1:]

            if int(date_str[0]) < 1911:
                date_str[0] = str(int(date_str[0])+1911)

            date_str[1] = date_str[1].zfill(2)
            date_str[2] = date_str[2].zfill(2)

            date_str = '/'.join(date_str)

            newdate = datetime.strptime(date_str, '%Y/%m/%d').strftime('%Y/%m/%d')
            return pd.Series([data_in[''] + ' (生日有誤)', newdate])

        except:
            try:
                if (date.today().year-65) < int(data_in['生日'][0:4]) <= date.today().year:
                    newdate = datetime.strptime(data_in['生日'], '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

                elif (date.today().year-65-1911) < int(data_in['生日'][0:2]) <= (date.today().year-1911):
                    newdate = datetime.strptime(str(int(data_in['生日']) + 1911*10**(len(str(data_in['生日']))-2)), '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

                elif (date.today().year-65-1911) < int(data_in['生日'][0:3]) <= (date.today().year-1911):
                    newdate = datetime.strptime(str(int(data_in['生日']) + 1911*10**(len(str(data_in['生日']))-3)), '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([data_in[''] + ' (生日有誤)', newdate])

            except:
                return pd.Series([data_in[''] + ' (生日有誤無法更正)', ''])

#------------------------------------------------------------------------------

def load_image(impath, image_size):

    return imtk.PhotoImage(im.open(impath).resize((image_size, image_size)))

#------------------------------------------------------------------------------

def add_char(add):

    with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
        word = f.read().split(',')

    for value in add.split(','):
        if value not in word:
            word.append(value)

    with open('file/word.txt', 'w', encoding='utf_8_sig') as f:
        f.write(','.join(word))

#------------------------------------------------------------------------------

def minus_char(minus):

    with open('file/word.txt', 'r', encoding='utf_8_sig') as f:
        word = f.read().split(',')

    for value in minus.split(','):
        if value in word:
            word.remove(value)

    with open('file/word.txt', 'w', encoding='utf_8_sig') as f:
        f.write(','.join(word))

#==============================================================================

HsuMei().run()