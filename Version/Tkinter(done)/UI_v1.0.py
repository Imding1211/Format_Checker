
from tkinter.scrolledtext import ScrolledText
from openpyxl import load_workbook
from datetime import datetime
from tkinter import filedialog
from tkinter import ttk
from tkinter import *
import customtkinter
import pandas as pd
import numpy as np

#==============================================================================

def run():

    try:
        path = filedialog.askopenfilename(title="選取檔案",multiple=False)

        wb = load_workbook(path)
        sheet = wb.active

        ID   = []
        name = []
        sex  = []
        date = []
        
        index = 2
        while True:
            if sheet.cell(row = index, column = 1).value == None:
                break
            ID.append(sheet.cell(row = index, column = 1).value)
            name.append(sheet.cell(row = index, column = 3).value)
            sex.append(sheet.cell(row = index, column = 4).value)
            date.append(sheet.cell(row = index, column = 5).value)
            index += 1
        
        dict = {'身分證字號': ID[1:], '姓名': name[1:], '性別': sex[1:], '生日': date[1:]}
        data = pd.DataFrame(dict).astype('str')
        data[''] = data.apply(chack_sapce, axis=1)
        data[''] = data.apply(chack_ID, axis=1)
        data[['', '更正後生日']] = data.apply(chack_date, axis=1)
        data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] = data.loc[data.duplicated(subset=['身分證字號'], keep=False), ''] + ' (身份證字號重複)'
        data = data.reindex(columns=['身分證字號', '姓名', '性別', '生日', '更正後生日', ''])

        savepath = path[0:-len(path.split('/')[-1])]
        savename = str(name_entry.get())
        if savename == '':
            savename = 'result.csv'
        elif savename[-4:] != '.csv':
            savename = savename + '.csv'
        data[data[''] != ''].to_csv(savepath+savename, index=False, encoding='utf_8_sig')

        text="資料有問題的數量:"+str(data[data[''] != ''].shape[0])+"\n"
        text_space.insert(END, text)
        text_space.see(END)

    except:
        text="檔案無法辨識"
        text_space.insert(END, text)
        text_space.see(END)

#------------------------------------------------------------------------------

def chack_ID(data_in):

    code_list = ['10','11','12','13','14','15','16','17','34','18','19','20','21','22','35','23','24','25','26','27','28','29','32','30','31','33']
    
    weight_list = [1,9,8,7,6,5,4,3,2,1,1]
    
    data_in['身分證字號'] = data_in['身分證字號'].strip()

    try:
        newcode = [int(x) for x in str(code_list[ord(data_in['身分證字號'][0])-65] + data_in['身分證字號'][1:])]

        if newcode[2] not in [1,2]:
            return str(data_in['']) + ' (外籍人士請留證件)'

        else:
            num = sum(np.multiply(newcode, weight_list))
        
            if num % 10 == 0:

                if data_in['身分證字號'][1] == '1' and data_in['性別'] != '男':
                    return str(data_in['']) + ' (性別有誤請留證件)'

                elif data_in['身分證字號'][1] == '2' and data_in['性別'] != '女':
                    return str(data_in['']) + ' (性別有誤請留證件)'

                else:
                    return str(data_in['']) + ''

            else:
                return str(data_in['']) + ' (身分證有誤請留證件)'
        
    except:
        if not data_in['身分證字號'][1].isnumeric():
            return str(data_in['']) + ' (外籍人士請留證件)'

        else:
            return str(data_in['']) + ' (身分證異常請留證件)'

#------------------------------------------------------------------------------

def chack_date(data_in):
    
    try:
        newdate = datetime.strptime(data_in['生日'], '%Y/%m/%d').strftime('%Y/%m/%d')
        
        if data_in['生日'] == newdate:
            return pd.Series([str(data_in['']), ''])
        
        else:
            return pd.Series([str(data_in['']) + ' (生日有誤)', newdate])
            
    except:
        count = 0
        date_str = ''
        
        for value in [x for x in data_in['生日']]:
            
            if value.isnumeric():
                date_str += value
                
            else:
                if count == 2:
                    break
                
                else:
                    date_str += '/'
                    count += 1

        try:
            newdate = datetime.strptime(date_str, '%Y/%m/%d').strftime('%Y/%m/%d')
            return pd.Series([str(data_in['']) + ' (生日有誤)', newdate])
            
        except:
            try:
                if date_str[0:4].isnumeric() and int(date_str[0:4]) <= 2022:
                    newdate = datetime.strptime(data_in['生日'], '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([str(data_in['']) + ' (生日有誤)', newdate])

                else:
                    date_str = ''
                    for value in [x for x in data_in['生日']]:
                        if value.isnumeric():
                            date_str += value

                    newdate = datetime.strptime(str(int(date_str) + 1911*10**(len(str(date_str))-2)), '%Y%m%d').strftime('%Y/%m/%d')
                    return pd.Series([str(data_in['']) + ' (生日有誤)', newdate])
            
            except:
                return pd.Series([str(data_in['']) + ' (生日有誤無法更正)', ''])

#------------------------------------------------------------------------------

def chack_sapce(data_in):
    
    if data_in['身分證字號'][0] == ' ' and data_in['身分證字號'][-1] == ' ':
        return '(身分證字號前後有空格)'
    
    elif data_in['身分證字號'][-1] == ' ':
        return '(身分證字號後有空格)'
    
    elif data_in['身分證字號'][0] == ' ':
        return '(身分證字號前有空格)'
    
    else:
        return ''

#==============================================================================

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("blue")
text_font = ('Courier New','10')

app = customtkinter.CTk()
app.geometry("500x350")
app.title("Hsu Mei")
app.resizable(0,0) 
app.iconbitmap('cat_icon.ico')
app.grid_columnconfigure(0, weight=1)

#------------------------------------------------------------------------------

frame_up = customtkinter.CTkFrame(master=app, corner_radius=10)
frame_up.grid(row=0, column=0, sticky=NSEW, pady=5, padx=5)
frame_up.grid_columnconfigure(0, weight=10)
frame_up.grid_columnconfigure(1, weight=1)

name_entry = customtkinter.CTkEntry(master=frame_up, 
                                    placeholder_text="存檔名稱",
                                    border_width=2,
                                    corner_radius=10)
name_entry.grid(row=0, column=0, sticky=EW, padx=10, pady=10)

select_button = customtkinter.CTkButton(master=frame_up,
                                        text="選取檔案",
                                        command=run)
select_button.grid(row=0, column=1, sticky=EW, padx=10, pady=10)

#------------------------------------------------------------------------------

frame_down = customtkinter.CTkFrame(master=app, corner_radius=10)
frame_down.grid(row=1, column=0, sticky=NSEW, pady=5, padx=5)
frame_down.grid_columnconfigure(0, weight=1)

text_space = ScrolledText(master=frame_down, 
                          height=16,  
                          font=text_font, 
                          wrap=WORD,
                          bd=0,
                          bg="#292929", 
                          fg="silver")
text_space.grid(row=0, column=0, padx=10, pady=10)

#------------------------------------------------------------------------------

app.mainloop()
